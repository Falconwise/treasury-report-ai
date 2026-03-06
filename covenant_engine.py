"""
Treasury Report AI - Covenant Compliance Engine
================================================
Core calculation engine for computing financial covenant ratios
and determining compliance status against bank-specific thresholds.

Designed for Saudi corporate treasury teams managing multi-bank
lending relationships (conventional + Islamic facilities).

Author: Treasury Report AI
"""

import pandas as pd
import openpyxl
from datetime import datetime, date
from dataclasses import dataclass, field
from typing import List, Dict, Optional
import re


# ══════════════════════════════════════════════════════
# DATA MODELS
# ══════════════════════════════════════════════════════

@dataclass
class FinancialData:
    """Parsed financial statement data (SAR Millions)."""
    # Income Statement
    revenue: float = 0.0
    cogs: float = 0.0
    gross_profit: float = 0.0
    sga: float = 0.0
    ebitda: float = 0.0
    depreciation: float = 0.0
    ebit: float = 0.0
    finance_charges: float = 0.0
    earnings_before_tax: float = 0.0
    tax: float = 0.0
    net_income: float = 0.0

    # Balance Sheet
    cash: float = 0.0
    total_current_assets: float = 0.0
    total_assets: float = 0.0
    total_current_liabilities: float = 0.0
    total_liabilities: float = 0.0
    total_equity: float = 0.0
    tangible_net_worth: float = 0.0
    long_term_borrowings: float = 0.0
    short_term_borrowings: float = 0.0


@dataclass
class Facility:
    """Bank credit facility details."""
    bank: str
    facility_type: str
    limit: float
    outstanding: float
    available: float
    utilization_pct: float
    rate: str
    maturity_date: str

    @property
    def remaining_years(self) -> float:
        """Calculate remaining years to maturity from reporting date."""
        try:
            for fmt in ["%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y"]:
                try:
                    mat = datetime.strptime(self.maturity_date, fmt).date()
                    break
                except ValueError:
                    continue
            else:
                return 3.0  # Default fallback

            reporting_date = date(2025, 12, 31)
            delta = (mat - reporting_date).days / 365.25
            return max(delta, 0.5)  # Floor at 0.5 years
        except Exception:
            return 3.0


@dataclass
class CovenantThreshold:
    """Bank-specific covenant threshold."""
    bank: str
    covenant_name: str
    metric: str
    threshold_str: str
    frequency: str

    @property
    def direction(self) -> str:
        """Whether the covenant is a minimum (>=) or maximum (<=)."""
        if ">=" in self.threshold_str:
            return "min"
        elif "<=" in self.threshold_str:
            return "max"
        return "min"

    @property
    def threshold_value(self) -> float:
        """Extract numeric threshold value."""
        match = re.search(r'SAR\s*(\d+)', self.threshold_str)
        if match:
            return float(match.group(1))
        match = re.search(r'([\d.]+)', self.threshold_str)
        if match:
            return float(match.group(1))
        return 0.0

    @property
    def is_absolute(self) -> bool:
        """Whether threshold is an absolute SAR amount (not a ratio)."""
        return "SAR" in self.threshold_str


@dataclass
class CovenantResult:
    """Result of a single covenant compliance check."""
    bank: str
    covenant_name: str
    metric: str
    threshold: float
    threshold_str: str
    actual_value: float
    actual_display: str
    direction: str  # "min" or "max"
    is_absolute: bool
    status: str  # "COMPLIANT", "WARNING", "BREACH"
    headroom_pct: float
    status_color: str  # "green", "amber", "red"


# ══════════════════════════════════════════════════════
# EXCEL READER
# ══════════════════════════════════════════════════════

class ExcelReader:
    """Reads and parses the input Excel workbook."""

    def __init__(self, filepath: str):
        self.filepath = filepath

    def read_financials(self) -> FinancialData:
        """Parse the Financials sheet into structured data."""
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        ws = wb["Financials"]

        data = FinancialData()
        mapping = {
            "Revenue": "revenue",
            "Cost of Goods Sold": "cogs",
            "Gross Profit": "gross_profit",
            "Selling, General & Administrative": "sga",
            "EBITDA": "ebitda",
            "Depreciation & Amortization": "depreciation",
            "EBIT (Operating Income)": "ebit",
            "Finance Charges": "finance_charges",
            "Earnings Before Tax": "earnings_before_tax",
            "Zakat & Income Tax": "tax",
            "Net Income": "net_income",
            "Cash & Cash Equivalents": "cash",
            "Total Current Assets": "total_current_assets",
            "Total Assets": "total_assets",
            "Total Current Liabilities": "total_current_liabilities",
            "Short-Term Borrowings": "short_term_borrowings",
            "Long-Term Borrowings": "long_term_borrowings",
            "Total Liabilities": "total_liabilities",
            "Total Equity": "total_equity",
            "Tangible Net Worth": "tangible_net_worth",
        }

        for row in ws.iter_rows(min_row=1, max_col=3, values_only=False):
            cell_val = row[0].value
            if cell_val and isinstance(cell_val, str):
                cell_val_stripped = cell_val.strip()
                if cell_val_stripped in mapping:
                    attr = mapping[cell_val_stripped]
                    val = row[1].value
                    if val is not None and isinstance(val, (int, float)):
                        setattr(data, attr, abs(float(val)) if attr in [
                            "cogs", "sga", "depreciation", "finance_charges", "tax"
                        ] else float(val))

        wb.close()
        return data

    def read_facilities(self) -> List[Facility]:
        """Parse the Facilities sheet."""
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        ws = wb["Facilities"]

        facilities = []
        for row in ws.iter_rows(min_row=5, max_col=8, values_only=True):
            if row[0] and row[0] != "TOTAL" and isinstance(row[2], (int, float)):
                util = row[5]
                if isinstance(util, str):
                    util = float(util.replace('%', ''))
                else:
                    util = float(util or 0)

                fac = Facility(
                    bank=str(row[0]),
                    facility_type=str(row[1]),
                    limit=float(row[2]),
                    outstanding=float(row[3]),
                    available=float(row[4]),
                    utilization_pct=util,
                    rate=str(row[6]),
                    maturity_date=str(row[7]),
                )
                facilities.append(fac)

        wb.close()
        return facilities

    def read_covenants(self) -> List[CovenantThreshold]:
        """Parse the Covenants sheet."""
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        ws = wb["Covenants"]

        covenants = []
        for row in ws.iter_rows(min_row=5, max_col=5, values_only=True):
            if row[0] and row[1]:
                cov = CovenantThreshold(
                    bank=str(row[0]),
                    covenant_name=str(row[1]),
                    metric=str(row[2]),
                    threshold_str=str(row[3]),
                    frequency=str(row[4]),
                )
                covenants.append(cov)

        wb.close()
        return covenants


# ══════════════════════════════════════════════════════
# COVENANT CALCULATOR
# ══════════════════════════════════════════════════════

class CovenantCalculator:
    """
    Calculates financial covenant ratios and determines compliance.
    
    WARNING_THRESHOLD: If actual value is within 15% of breach,
    status is WARNING (amber). Beyond that = BREACH (red).
    """

    WARNING_BUFFER = 0.15  # 15% headroom triggers warning

    def __init__(self, financials: FinancialData, facilities: List[Facility]):
        self.fin = financials
        self.facilities = facilities
        self._total_outstanding = sum(f.outstanding for f in facilities)
        self._total_annual_principal = self._calc_total_annual_principal()

    def _calc_total_annual_principal(self) -> float:
        """Calculate total annual principal repayments across all facilities."""
        total = 0.0
        for fac in self.facilities:
            years = fac.remaining_years
            annual_principal = fac.outstanding / years
            total += annual_principal
        return total

    def _calc_annual_principal_for_bank(self, bank: str) -> float:
        """Calculate annual principal for a specific bank's facilities."""
        total = 0.0
        for fac in self.facilities:
            if fac.bank == bank:
                years = fac.remaining_years
                total += fac.outstanding / years
        return total

    def calc_dscr(self, bank: str = None) -> float:
        """
        Debt Service Coverage Ratio.
        Formula: (EBITDA - Tax) / (Annual Principal Repayments + Finance Charges)
        """
        numerator = self.fin.ebitda - self.fin.tax
        if bank:
            principal = self._calc_annual_principal_for_bank(bank)
            # Pro-rata finance charges based on outstanding
            bank_outstanding = sum(f.outstanding for f in self.facilities if f.bank == bank)
            if self._total_outstanding > 0:
                finance_share = self.fin.finance_charges * (bank_outstanding / self._total_outstanding)
            else:
                finance_share = self.fin.finance_charges
            denominator = principal + finance_share
        else:
            denominator = self._total_annual_principal + self.fin.finance_charges

        return round(numerator / denominator, 2) if denominator > 0 else 999.0

    def calc_debt_to_equity(self) -> float:
        """Debt-to-Equity Ratio = Total Liabilities / Total Equity."""
        if self.fin.total_equity > 0:
            return round(self.fin.total_liabilities / self.fin.total_equity, 2)
        return 999.0

    def calc_current_ratio(self) -> float:
        """Current Ratio = Current Assets / Current Liabilities."""
        if self.fin.total_current_liabilities > 0:
            return round(self.fin.total_current_assets / self.fin.total_current_liabilities, 2)
        return 999.0

    def calc_icr(self) -> float:
        """Interest Coverage Ratio = EBIT / Finance Charges."""
        if self.fin.finance_charges > 0:
            return round(self.fin.ebit / self.fin.finance_charges, 2)
        return 999.0

    def calc_debt_to_ebitda(self) -> float:
        """Leverage Ratio = Total Debt / EBITDA."""
        total_debt = self.fin.short_term_borrowings + self.fin.long_term_borrowings
        if self.fin.ebitda > 0:
            return round(total_debt / self.fin.ebitda, 2)
        return 999.0

    def calc_net_worth(self) -> float:
        """Return Tangible Net Worth (SAR M)."""
        return self.fin.tangible_net_worth

    def _get_ratio_value(self, metric: str, bank: str) -> float:
        """Get the calculated ratio value for a given metric."""
        metric_lower = metric.lower()
        if "dscr" in metric_lower:
            return self.calc_dscr(bank=bank)
        elif "debt/equity" in metric_lower or "debt-to-equity" in metric_lower:
            return self.calc_debt_to_equity()
        elif "current" in metric_lower:
            return self.calc_current_ratio()
        elif "icr" in metric_lower:
            return self.calc_icr()
        elif "debt/ebitda" in metric_lower or "leverage" in metric_lower:
            return self.calc_debt_to_ebitda()
        elif "net worth" in metric_lower or "worth" in metric_lower:
            return self.calc_net_worth()
        return 0.0

    def _determine_status(self, actual: float, threshold: float, direction: str) -> tuple:
        """
        Determine compliance status and headroom.
        Returns: (status, headroom_pct, color)
        """
        if direction == "min":
            # Must be >= threshold
            if actual >= threshold:
                headroom = ((actual - threshold) / threshold) * 100
                if headroom <= self.WARNING_BUFFER * 100:
                    return "WARNING", round(headroom, 1), "amber"
                return "COMPLIANT", round(headroom, 1), "green"
            else:
                headroom = ((actual - threshold) / threshold) * 100
                return "BREACH", round(headroom, 1), "red"
        else:
            # Must be <= threshold (max)
            if actual <= threshold:
                headroom = ((threshold - actual) / threshold) * 100
                if headroom <= self.WARNING_BUFFER * 100:
                    return "WARNING", round(headroom, 1), "amber"
                return "COMPLIANT", round(headroom, 1), "green"
            else:
                headroom = ((threshold - actual) / threshold) * 100
                return "BREACH", round(headroom, 1), "red"

    def evaluate_covenants(self, covenants: List[CovenantThreshold]) -> List[CovenantResult]:
        """Evaluate all covenants and return compliance results."""
        results = []

        for cov in covenants:
            actual = self._get_ratio_value(cov.metric, cov.bank)
            threshold = cov.threshold_value
            direction = cov.direction
            is_abs = cov.is_absolute

            status, headroom, color = self._determine_status(actual, threshold, direction)

            if is_abs:
                actual_display = f"SAR {actual:.0f}M"
            else:
                actual_display = f"{actual:.2f}x"

            result = CovenantResult(
                bank=cov.bank,
                covenant_name=cov.covenant_name,
                metric=cov.metric,
                threshold=threshold,
                threshold_str=cov.threshold_str,
                actual_value=actual,
                actual_display=actual_display,
                direction=direction,
                is_absolute=is_abs,
                status=status,
                headroom_pct=headroom,
                status_color=color,
            )
            results.append(result)

        return results


# ══════════════════════════════════════════════════════
# CONVENIENCE FUNCTION
# ══════════════════════════════════════════════════════

def run_covenant_analysis(excel_path: str) -> dict:
    """
    Complete covenant analysis pipeline.
    
    Returns dict with:
        - financials: FinancialData
        - facilities: List[Facility]
        - covenants: List[CovenantThreshold]
        - results: List[CovenantResult]
        - summary: dict with counts
    """
    reader = ExcelReader(excel_path)
    financials = reader.read_financials()
    facilities = reader.read_facilities()
    covenants = reader.read_covenants()

    calculator = CovenantCalculator(financials, facilities)
    results = calculator.evaluate_covenants(covenants)

    compliant = sum(1 for r in results if r.status == "COMPLIANT")
    warning = sum(1 for r in results if r.status == "WARNING")
    breach = sum(1 for r in results if r.status == "BREACH")

    summary = {
        "total_covenants": len(results),
        "compliant": compliant,
        "warning": warning,
        "breach": breach,
        "overall_status": "BREACH" if breach > 0 else ("WARNING" if warning > 0 else "COMPLIANT"),
        "reporting_date": "31 December 2025",
        "company_name": "Al Watan Industries",
    }

    return {
        "financials": financials,
        "facilities": facilities,
        "covenants": covenants,
        "results": results,
        "summary": summary,
    }


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "sample_data.xlsx"
    analysis = run_covenant_analysis(path)

    print("\n" + "=" * 70)
    print("  TREASURY REPORT AI - Covenant Compliance Summary")
    print("=" * 70)
    print(f"  Company: {analysis['summary']['company_name']}")
    print(f"  Period:  {analysis['summary']['reporting_date']}")
    print(f"  Overall: {analysis['summary']['overall_status']}")
    print("-" * 70)

    for r in analysis["results"]:
        icon = "+" if r.status == "COMPLIANT" else ("!" if r.status == "WARNING" else "X")
        print(f"  {icon} {r.bank:30s} | {r.covenant_name:30s} | "
              f"{r.actual_display:>12s} vs {r.threshold_str:>12s} | "
              f"{r.status:10s} | Headroom: {r.headroom_pct:+.1f}%")

    print("=" * 70)